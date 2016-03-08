# -*- coding: utf-8 -*-

from openpyxl import Workbook

from scraperPurchase import *


class CreateExcelReport:
    def __init__(self):
        self.dbs = DBSaver()

    def __del__(self):
        pass

    def writeExcel(self):
        self.dbs.postETL()
        cur = self.dbs.conn.cursor()
        wb = Workbook()

        rc = self.writeGagarinskiySheet(cur, wb)
        print "writeMainSheet: DONE", rc

        for tag in [u'Гагаринский', u'ВоробьевыГоры', u'Университетский', u'Вернадского', u'Ленинский',
                    u'Академический', u'ПрефектураЮЗАО', u'Благоустройство']:
            rc = self.writeTagWinnersSheet(cur, wb, tag)
            print "writeTagWinnersSheet(", tag.encode("utf-8"), ") : DONE", rc

        # rc = self.writeMoscowWinnersSheet(cur, wb)
        # print "writeMoscowWinnersSheet: DONE", rc

        # rc = self.writeAkademicheskiyWinnersSheet(cur, wb)
        # print "writeAkademicheskiyWinnersSheet: DONE", rc
        #
        # rc = self.writePrefekturaUSAOWinnersSheet(cur, wb)
        # print "writePrefekturaUSAOWinnersSheet: DONE", rc

        rc = self.writeMetaSheets(cur, wb)

        # Save the file
        wb.save("purchases.xlsx")
        cur.close()
        print "saved: purchases.xlsx"

    def writeMetaSheets(self, cur, wb):
        self.writeSQLSheet(cur, wb, "Scraping Errors", """
SELECT message,
       exc_type,
       exc_value,
       exc_traceback,
       max(loaddate) AS max_date,
       count(1) AS numErrors
FROM tErrorLog
WHERE loadDate > now()- interval '10 days'
GROUP BY message,
         exc_type,
         exc_value,
         exc_traceback
ORDER BY numErrors DESC,
         max_date DESC
		""")
        print "write Scraping Errors: done"

        self.writeSQLSheet(cur, wb, "Last1000Errors", """SELECT * FROM tErrorLog  ORDER BY loadDate desc LIMIT 1000""")
        print "write Last1000Errors: done"

        self.writeSQLSheet(cur, wb, "ProxyStats", """
with httpRes as (
	select 	proxy,
		to_char(_loadDate, 'YYYY-MM-DD') as dt,
		CASE
			WHEN result like 'Success%' THEN True
			ELSE FALSE
		END as Suceeded
	from tHTTPProxyResult
	order by _loadDate desc
), htGrp as (
	select proxy, dt, count(proxy) as succ , 0 as fail
	from httpRes where Suceeded group by proxy, dt
	union all
	select proxy, dt, 0 as succ , count(proxy) as fail
	from httpRes where NOT Suceeded group by proxy, dt
)
select proxy, dt, sum(succ) as succ, sum(fail) as fails, sum(succ)/(sum(succ) + sum(fail)) as successRate
from htGrp
group by proxy, dt
""")
        print "write ProxyStats: done"

        self.writeSQLSheet(cur, wb, "PotentialBIDs", """ select * from tPurchaseContracts tpc where winnerINN is not null
and not exists (select 1 from tPurchaseBid           tpb where tpc.purchaseId=tpb.purchaseId)
and lower(customerName) like '%моск%' order by price DESC LIMIT 500 """)
        print "write PotentialBIDs: done"

    def writeSQLSheet(self, cur, wb, sheetName, sql):
        rowCount = 0
        ws = wb.create_sheet()
        ws.title = sheetName
        cur.execute(sql)
        rows = cur.fetchall()
        ws.append([desc[0] for desc in cur.description])
        for row in rows:
            # Rows can also be appended
            ws.append(row)
            rowCount += 1

        ws.auto_filter.ref = "A:Z"
        # ws.auto_filter.add_filter_column(0, ['Fatal*'], False)
        return rowCount

    def writeGagarinskiySheet(self, cur, wb):
        rowCount = 0
        ws = wb.active  # create_sheet()
        ws.title = "Gagarinskiy"
        cur.execute("""
SELECT pp.purchaseId,
       orderId,
       customerName,
       title,
       purchaseType,
       stage ,
       contractAmount,
       _url,
       to_char( coalesce(submitFinish, submitStart, requestPublished), 'YYYY-mm') as SomeDate,
       _loaddate,
       requestPublished,
       submitStart,
       submitFinish,
       responsible,
       contractAmountT,
       requestPublishedT,
       submitStartT,
       submitFinishT
FROM tPurchase pp
JOIN tPurchaseDetails ppd ON pp.purchaseId = ppd.purchaseId
WHERE ppd.title IS NOT NULL
and ppd.purchaseId in (select purchaseId from tPurchaseTags WHERE tagLabel in ('Гагаринский', 'ВоробьевыГоры', 'Университетский'))
		""")
        # select * from vArtList order by loadDate desc
        rows = cur.fetchall()
        ws.append([desc[0] for desc in cur.description])
        for row in rows:
            # Rows can also be appended
            ws.append(row)
            rowCount += 1
        ws.auto_filter.ref = "A:Z"
        return rowCount

    def writeTagWinnersSheet(self, cur, wb, tagName):
        rowCount = 0
        ws = wb.create_sheet()
        ws.title = tagName
        cur.execute("""
SELECT
       ppd.customerName,
       title,
       winnerName,
       winnerINN,
       purchaseType,
       stage ,
       contractAmount as Tender_Amount,
       price as Contract_Price,
       priceT,
       _url,
       to_char( coalesce(submitFinish, submitStart, requestPublished), 'YYYY-mm') as SomeDate,
       contractStatus,
       pp._loaddate,
       requestPublished,
       submitStart,
       submitFinish,
       responsible,
       pp.purchaseId,
       orderId
FROM tPurchase pp
JOIN tPurchaseDetails ppd ON pp.purchaseId = ppd.purchaseId
LEFT JOIN tPurchaseContracts pcc on  pp.purchaseId = pcc.purchaseId
WHERE ppd.title IS NOT NULL
and ppd.purchaseId in (select purchaseId from tPurchaseTags WHERE tagLabel in (%s))
order by Tender_Amount desc
		""", [tagName])
        rows = cur.fetchall()
        ws.append([desc[0] for desc in cur.description])
        for row in rows:
            # Rows can also be appended
            ws.append(row)
            rowCount += 1

        ws.auto_filter.ref = "A:Z"
        # ws.auto_filter.add_filter_column(0, ['Fatal*'], False)
        return rowCount

    def writeMoscowWinnersSheet(self, cur, wb):
        rowCount = 0
        ws = wb.create_sheet()
        ws.title = "Moscow Winners"
        cur.execute("""
SELECT
       ppd.customerName,
       title,
       winnerName,
       winnerInn,
       purchaseType,
       stage ,
       contractAmount as Tender_Amount,
       price as Contract_Price,
       _url,
       to_char( coalesce(submitFinish, submitStart, requestPublished), 'YYYY-mm') as SomeDate,
       pp._loaddate,
       requestPublished,
       submitStart,
       submitFinish,
       responsible,
       pp.purchaseId,
       orderId,
       priceT as Contract_Price_Text
FROM tPurchase pp
JOIN tPurchaseDetails ppd ON pp.purchaseId = ppd.purchaseId
LEFT JOIN tPurchaseContracts pcc on  pp.purchaseId = pcc.purchaseId
WHERE lower(ppd.title) like '%москв%' OR lower(ppd.customerName) like '%москв%'
order by Tender_Amount desc
		""")
        rows = cur.fetchall()
        ws.append([desc[0] for desc in cur.description])
        for row in rows:
            # Rows can also be appended
            ws.append(row)
            rowCount += 1

        ws.auto_filter.ref = "A:Z"
        # ws.auto_filter.add_filter_column(0, ['Fatal*'], False)
        return rowCount


a = CreateExcelReport()
a.writeExcel()
